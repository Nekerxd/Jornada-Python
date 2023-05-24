from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.image import Image

class LeitorAcoes:
    def __init__(self, caminho_arquivo: str = ""):
        self.caminho_arquivo = caminho_arquivo
        self.dados = []
    
    def processa_arquivo(self,acao: str):
        with open(f'{self.caminho_arquivo}{acao}.txt', 'r') as arquivo_cotacao:
            linhas = arquivo_cotacao.readlines()
            self.dados = [linha.replace("\n", "").split(";") for linha in linhas]

class PropriedadeSerieGrafico:
    def __init__(self, espessura: int, cor_preenchimento: str):
        self.espessura = espessura
        self.cor_preenchimento = cor_preenchimento


class GerenciadorPlanilha:
    def __init__(self):
        self.workbook = Workbook()
        self.planilha_ativa = None
    
    
    def adiciona_planilha(self, titulo_planilha: str = ""):
        nova_planilha = self.workbook.create_sheet(titulo_planilha)
        self.workbook.active = nova_planilha
        self.planilha_ativa = nova_planilha
        
        return nova_planilha
    
    
    def adiciona_linha(self, dados: list):
        self.planilha_ativa.append(dados)
        
    
    def atualiza_celula(self, celula: str, dado):
        self.planilha_ativa[celula] = dado
        
    
    def mescla_celulas(self, celula_inicio: str, celula_fim: str):
        self.planilha_ativa.merge_cells(f'{celula_inicio}:{celula_fim}')
    
    def aplica_estilos(self, celula: str, estilos: list):
        for estilo in estilos:
            setattr(self.planilha_ativa[celula], estilo[0], estilo[1])

    def adiciona_grafico_linha(self, celula: str, comprimento: float, altura: float, titulo: str, titulo_eixo_x: str, titulo_eixo_y: str, referencia_eixo_x: Reference, referencia_eixo_y: Reference, propriedades_grafico: list):
        
        grafico = LineChart()
        grafico.width = comprimento
        grafico.height = altura
        grafico.title = titulo
        grafico.x_axis.title = titulo_eixo_x
        grafico.y_axis.title = titulo_eixo_y
        
        grafico.add_data(referencia_eixo_x)
        grafico.set_categories(referencia_eixo_y)
        
        for serie, propriedade in zip(grafico.series, propriedades_grafico):
            serie.graphicalProperties.line.width = propriedade.espessura
            serie.graphicalProperties.line.solidFill = propriedade.cor_preenchimento
        
        self.planilha_ativa.add_chart(grafico, celula)
    
    
    def adiciona_imagem(self, celula: str, caminho_imagem: str):
        imagem = Image(caminho_imagem)
        self.planilha_ativa.add_image(imagem, celula)
    
    
    def salva_arquivo(self, caminho_arquivo: str):
        self.workbook.save(caminho_arquivo)